#!/usr/bin/env python3
"""
Script para sincronizar datos entre Excel y la base de datos
"""

import openpyxl
from database import CasamientoDatabase
import sys

EXCEL_PATH = '/Users/acardoso/Projects/Planificacion_Casamiento_20260218.xlsx'

def importar_desde_excel():
    """Importar datos del Excel a la base de datos"""
    print("📥 Importando datos desde Excel...")
    
    db = CasamientoDatabase()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Importar presupuesto
    print("\n💰 Importando presupuesto...")
    ws_presupuesto = wb['Presupuesto']
    
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Limpiar tabla de presupuesto
    cursor.execute('DELETE FROM presupuesto')
    
    for row in range(2, 32):  # Filas 2-31 tienen los items
        categoria = ws_presupuesto.cell(row=row, column=1).value
        item = ws_presupuesto.cell(row=row, column=2).value
        
        if categoria and item:
            estimado = ws_presupuesto.cell(row=row, column=5).value or 0
            real = ws_presupuesto.cell(row=row, column=6).value or 0
            pagado = ws_presupuesto.cell(row=row, column=7).value or 0
            proveedor = ws_presupuesto.cell(row=row, column=3).value or ''
            
            cursor.execute('''
                INSERT INTO presupuesto (categoria, item, estimado, real, pagado, proveedor)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (categoria, item, estimado, real, pagado, proveedor))
            
            print(f"  ✓ {categoria} - {item}: ${estimado:,.0f}")
    
    conn.commit()
    print(f"  ✅ {cursor.rowcount} items de presupuesto importados")
    
    # Importar invitados
    print("\n👥 Importando invitados...")
    ws_invitados = wb['Invitados']
    
    # Limpiar tabla de invitados
    cursor.execute('DELETE FROM invitados')
    
    invitados_count = 0
    for row in range(4, 54):  # Filas 4-53 (50 invitados)
        nombre = ws_invitados.cell(row=row, column=2).value
        
        if nombre and nombre.strip():
            telefono = ws_invitados.cell(row=row, column=3).value or ''
            email = ws_invitados.cell(row=row, column=4).value or ''
            grupo = ws_invitados.cell(row=row, column=5).value or ''
            invitacion = ws_invitados.cell(row=row, column=6).value == 'Sí'
            confirmacion = ws_invitados.cell(row=row, column=7).value or 'Pendiente'
            menu = ws_invitados.cell(row=row, column=9).value or ''
            alergias = ws_invitados.cell(row=row, column=10).value or ''
            mesa = ws_invitados.cell(row=row, column=11).value
            
            cursor.execute('''
                INSERT INTO invitados (nombre, telefono, email, grupo, invitacion_enviada,
                                      confirmacion, menu, alergias, mesa)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (nombre, telefono, email, grupo, invitacion, confirmacion, menu, alergias, mesa))
            
            invitados_count += 1
            print(f"  ✓ {nombre} ({grupo})")
    
    conn.commit()
    print(f"  ✅ {invitados_count} invitados importados")
    
    # Importar tareas
    print("\n📅 Importando tareas...")
    ws_cronograma = wb['Cronograma']
    
    # Limpiar tabla de tareas
    cursor.execute('DELETE FROM tareas')
    
    tareas_count = 0
    for row in range(7, 28):  # Filas con tareas
        fecha = ws_cronograma.cell(row=row, column=1).value
        tarea = ws_cronograma.cell(row=row, column=2).value
        
        if fecha and tarea:
            categoria = ws_cronograma.cell(row=row, column=3).value or ''
            responsable = ws_cronograma.cell(row=row, column=4).value or ''
            estado = ws_cronograma.cell(row=row, column=5).value or 'Pendiente'
            prioridad = ws_cronograma.cell(row=row, column=6).value or 'MEDIA'
            
            cursor.execute('''
                INSERT INTO tareas (fecha, tarea, categoria, responsable, estado, prioridad)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (str(fecha), tarea, categoria, responsable, estado, prioridad))
            
            tareas_count += 1
            print(f"  ✓ {fecha} - {tarea}")
    
    conn.commit()
    print(f"  ✅ {tareas_count} tareas importadas")
    
    # Importar proveedores
    print("\n🏢 Importando proveedores...")
    ws_proveedores = wb['Proveedores']
    
    # Limpiar tabla de proveedores
    cursor.execute('DELETE FROM proveedores')
    
    proveedores_count = 0
    for row in range(4, 50):  # Filas con proveedores
        categoria = ws_proveedores.cell(row=row, column=1).value
        nombre = ws_proveedores.cell(row=row, column=2).value
        
        if categoria and nombre:
            contacto = ws_proveedores.cell(row=row, column=3).value or ''
            telefono = ws_proveedores.cell(row=row, column=4).value or ''
            email = ws_proveedores.cell(row=row, column=5).value or ''
            direccion = ws_proveedores.cell(row=row, column=6).value or ''
            precio = ws_proveedores.cell(row=row, column=7).value or ''
            contratado = ws_proveedores.cell(row=row, column=8).value == 'Sí'
            notas = ws_proveedores.cell(row=row, column=9).value or ''
            
            cursor.execute('''
                INSERT INTO proveedores (categoria, nombre, contacto, telefono, email,
                                        direccion, precio, contratado, notas)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (categoria, nombre, contacto, telefono, email, direccion, precio, contratado, notas))
            
            proveedores_count += 1
            print(f"  ✓ {categoria} - {nombre}")
    
    conn.commit()
    print(f"  ✅ {proveedores_count} proveedores importados")
    
    conn.close()
    wb.close()
    
    print("\n✅ Importación completada exitosamente!")
    print(f"\n📊 Resumen:")
    print(f"   - Presupuesto: {cursor.rowcount} items")
    print(f"   - Invitados: {invitados_count} personas")
    print(f"   - Tareas: {tareas_count} tareas")
    print(f"   - Proveedores: {proveedores_count} proveedores")

def exportar_a_excel():
    """Exportar datos de la base de datos al Excel"""
    print("📤 Exportando datos a Excel...")
    
    db = CasamientoDatabase()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Exportar presupuesto
    print("\n💰 Exportando presupuesto...")
    ws_presupuesto = wb['Presupuesto']
    
    presupuesto = db.get_presupuesto()
    row = 2
    
    for categoria_data in presupuesto:
        for item in categoria_data['items']:
            ws_presupuesto.cell(row=row, column=1, value=categoria_data['nombre'])
            ws_presupuesto.cell(row=row, column=2, value=item['item'])
            ws_presupuesto.cell(row=row, column=3, value=item['proveedor'])
            ws_presupuesto.cell(row=row, column=5, value=item['estimado'])
            ws_presupuesto.cell(row=row, column=6, value=item['real'])
            ws_presupuesto.cell(row=row, column=7, value=item['pagado'])
            row += 1
    
    print(f"  ✅ Presupuesto exportado")
    
    # Exportar invitados
    print("\n👥 Exportando invitados...")
    ws_invitados = wb['Invitados']
    
    invitados = db.get_invitados()
    row = 4
    
    for inv in invitados:
        ws_invitados.cell(row=row, column=1, value=row-3)  # Número
        ws_invitados.cell(row=row, column=2, value=inv['nombre'])
        ws_invitados.cell(row=row, column=3, value=inv['telefono'])
        ws_invitados.cell(row=row, column=4, value=inv['email'])
        ws_invitados.cell(row=row, column=5, value=inv['grupo'])
        ws_invitados.cell(row=row, column=6, value='Sí' if inv['invitacion_enviada'] else 'No')
        ws_invitados.cell(row=row, column=7, value=inv['confirmacion'])
        ws_invitados.cell(row=row, column=9, value=inv['menu'])
        ws_invitados.cell(row=row, column=10, value=inv['alergias'])
        ws_invitados.cell(row=row, column=11, value=inv['mesa'])
        row += 1
    
    print(f"  ✅ {len(invitados)} invitados exportados")
    
    # Exportar tareas
    print("\n📅 Exportando tareas...")
    ws_cronograma = wb['Cronograma']
    
    tareas = db.get_tareas()
    row = 7
    
    for tarea in tareas:
        ws_cronograma.cell(row=row, column=1, value=tarea['fecha'])
        ws_cronograma.cell(row=row, column=2, value=tarea['tarea'])
        ws_cronograma.cell(row=row, column=3, value=tarea['categoria'])
        ws_cronograma.cell(row=row, column=4, value=tarea['responsable'])
        ws_cronograma.cell(row=row, column=5, value=tarea['estado'])
        ws_cronograma.cell(row=row, column=6, value=tarea['prioridad'])
        row += 1
    
    print(f"  ✅ {len(tareas)} tareas exportadas")
    
    # Guardar Excel
    wb.save(EXCEL_PATH)
    wb.close()
    
    print("\n✅ Exportación completada exitosamente!")
    print(f"   Archivo: {EXCEL_PATH}")

def sincronizar():
    """Sincronización bidireccional"""
    print("🔄 Sincronizando Excel ↔ Base de Datos...")
    print("\n1. Importando desde Excel...")
    importar_desde_excel()
    print("\n2. Los cambios en la web se guardarán automáticamente en la BD")
    print("3. Para exportar cambios de vuelta al Excel, ejecutá:")
    print("   python3 sync_excel.py exportar")

if __name__ == '__main__':
    if len(sys.argv) > 1:
        comando = sys.argv[1]
        if comando == 'importar':
            importar_desde_excel()
        elif comando == 'exportar':
            exportar_a_excel()
        elif comando == 'sync':
            sincronizar()
        else:
            print("Comandos disponibles:")
            print("  python3 sync_excel.py importar  - Importar desde Excel")
            print("  python3 sync_excel.py exportar  - Exportar a Excel")
            print("  python3 sync_excel.py sync      - Sincronizar")
    else:
        sincronizar()
